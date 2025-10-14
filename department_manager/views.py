import json
import os
import openpyxl
from django.shortcuts import get_object_or_404, render
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render,redirect
from datetime import datetime, time, timedelta
from registration.models import Profile
from core.utils import *
from poll.models import Request, Poll, Fields, RequestAnswer, RequestRecord, Fields
from manuals.models import Manuals
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from brigade.models import Brigade
from incident.models import Incident
from django.conf import settings
import xlwt

#Funcion para Encuestas Enviadas
@login_required
def department_main(request):
    session = int(check_profile_department(request))
    if session == 0:
        messages.error(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    manuals_list = Manuals.objects.filter(manual_name='Manual Departamento').first
    request_open_list_array = Request.objects.filter(request_state='Iniciada').order_by('-id')
    request_open_list = []
    for re in request_open_list_array:
        poll_data =  Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        request_territorial_name = re.user.first_name+' '+re.user.last_name
        request_open_list.append({'id':re.id,'request_name':re.request_name,'request_date':re.request_date,'request_type':request_type,'request_territorial_name':request_territorial_name})

    template_name = 'department_manager/department_main.html'  
    return render(request, template_name, {'request_open_list': request_open_list, 'manuals_list': manuals_list})

#Funcion para ver lista de Solicitudes Derivadas
@login_required
def department_list_derived(request,page=None):  
    session = int(check_profile_department(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    manuals_list = Manuals.objects.filter(manual_name='Manual Dirección').first
    request_delivery_list_array = Request.objects.filter(request_state='Derivada').order_by('-request_delivery')
    request_delivery_list = []
    for re in request_delivery_list_array:
        poll_data =  Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        #Cuadrialla a cargo
        user_brigade_incident_incharge = re.brigade.user.first_name+' '+re.brigade.user.last_name
        request_delivery_list.append({'id':re.id,'request_name':re.request_name,'request_delivery':re.request_delivery,'request_type':request_type,'user_brigade_incident_incharge':user_brigade_incident_incharge})    
    template_name = 'department_manager/department_list_derived.html'
    return render(request,template_name, {'request_delivery_list': request_delivery_list,'manuals_list': manuals_list})

#Funcion para ver lista de Solicitudes En Proceso
@login_required
def department_in_progress(request,page=None):  
    session = int(check_profile_department(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    manuals_list = Manuals.objects.filter(manual_name='Manual Dirección').first
    request_in_progress_list_array = Request.objects.filter(request_state='Proceso').order_by('-request_delivery')
    request_in_progress_list = []
    for re in request_in_progress_list_array:
        poll_data =  Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        #Cuadrialla a cargo
        user_brigade_incident_incharge = re.brigade.user.first_name+' '+re.brigade.user.last_name
        request_in_progress_list.append({'id':re.id,'request_name':re.request_name,'request_delivery':re.request_delivery,'request_accept':re.request_accept,'request_type':request_type,'user_brigade_incident_incharge':user_brigade_incident_incharge})    
    template_name = 'department_manager/department_in_progress.html'  
    return render(request,template_name, {'request_in_progress_list': request_in_progress_list,'manuals_list': manuals_list})

#Funcion para ver lista de Solicitudes Finalizadas
@login_required
def department_finish(request,page=None):  
    session = int(check_profile_department(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    manuals_list = Manuals.objects.filter(manual_name='Manual Dirección').first
    request_finish_list_array = Request.objects.filter(request_state='Finalizada').order_by('-request_delivery')
    request_in_finish_list = []
    for re in request_finish_list_array:
        poll_data =  Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        #Cuadrialla a cargo
        user_brigade_incident_incharge = re.brigade.user.first_name+' '+re.brigade.user.last_name
        request_in_finish_list.append({'id':re.id,'request_name':re.request_name,'request_delivery':re.request_delivery,'request_accept':re.request_accept,'request_accept':re.request_accept,'request_type':request_type,'user_brigade_incident_incharge':user_brigade_incident_incharge})    
    template_name = 'department_manager/department_finish.html'  
    return render(request,template_name, {'request_in_finish_list': request_in_finish_list,'manuals_list': manuals_list})

#Funcion para ver lista de Solicitudes Cerradas
@login_required
def department_list_closed(request,page=None):  
    session = int(check_profile_department(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    manuals_list = Manuals.objects.filter(manual_name='Manual Dirección').first
    request_close_list_array = Request.objects.filter(request_state='Cerrada').order_by('-request_delivery')
    request_closed_list = []
    for re in request_close_list_array:
        poll_data =  Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        #Cuadrialla a cargo
        user_brigade_incident_incharge = re.brigade.user.first_name+' '+re.brigade.user.last_name
        request_closed_list.append({'id':re.id,'request_name':re.request_name,'request_delivery':re.request_delivery,'request_accept':re.request_accept,'request_accept':re.request_accept,'request_close':re.request_close,'request_type':request_type,'user_brigade_incident_incharge':user_brigade_incident_incharge})    
    template_name = 'department_manager/department_list_closed.html'  
    return render(request,template_name, {'request_closed_list': request_closed_list,'manuals_list': manuals_list})

#Funcion para ver Encuestas Enviadas
@login_required
def department_view(request, request_id):
    session = int(check_profile_department(request))
    if session == 0:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    
    flow = type_flow(request)
    try:
        poll_request = Request.objects.get(pk=request_id)
        poll_id = poll_request.poll_id
        poll_request_state = poll_request.request_state
    except Request.DoesNotExist:
        messages.add_message(request, messages.INFO, 'Hubo un error al acceder a la solicitud')
        return redirect('check_group_main')
    
    # Obtener los campos estándar activos
    poll_fields_standard = Fields.objects.filter(poll_id=poll_id, state='Activo', kind_field='standard').order_by('id')
    poll_fields_other = Fields.objects.filter(poll_id=poll_id, state='bloqueado', kind_field='other').order_by('id')
    
    # Verificar que existen campos para el poll_id dado
    if not poll_fields_standard.exists():
        messages.add_message(request, messages.INFO, 'No se encontraron campos para la encuesta proporcionada')
        return redirect('check_group_main')
    
    # Obtener la encuesta creada
    poll_data = Poll.objects.get(pk=poll_id)
    
    # Obtener las respuestas de las solicitudes que coincidan con los campos rescatados
    request_answers = RequestAnswer.objects.filter(fields__in=poll_fields_standard).filter(request_id=request_id).select_related('fields', 'request', 'user')
    
    # Crear una lista de tuplas (nombre del campo, valor de respuesta)
    field_answer_pairs = []
    images = []
    lat = ''
    lon = ''
    incidence_video = ''
    incidence_audio = ''
    for field in poll_fields_standard:
        answer = request_answers.filter(fields=field).first()
        #field_answer_pairs.append((field.name, answer.request_answer_text if answer else ''))
        field_answer_pairs.append((field.name, answer.request_answer_text if answer else ''))
        if field.name == 'incidence_latitud':
            lat = answer.request_answer_text if answer else ''
        if field.name == 'incidence_longitud':
            lon = answer.request_answer_text if answer else ''
        if field.name == 'incidence_video':
            incidence_video = answer.request_answer_text if answer else ''
        if field.name == 'incidence_audio':
            incidence_audio = answer.request_answer_text if answer else ''

    for field in poll_fields_standard:
        if field.name == 'incidence_image':
            answer = request_answers.filter(fields=field)
            for a in answer:
                incidence_image = a.request_answer_text if answer else ''
                images.append(incidence_image)

            
    brigades = Brigade.objects.select_related('user').filter(state='Activa')
    #usuario que genera la incidencia
    user_create_incident = poll_request.user.first_name+' '+poll_request.user.last_name

    # Preparar la lista con los datos necesarios
    brigades_data = []
    for brigade in brigades:
        user = brigade.user
        brigades_data.append({
            'brigade_id': brigade.id,
            'user_id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name
        })
    fields_record= RequestRecord.objects.filter(request_id=request_id)
    template_name = 'department_manager/department_view.html'
    context = {
        'brigades_data': brigades_data,
        'request_id': request_id,
        'poll_data': poll_data,
        'field_answer_pairs': field_answer_pairs,
        'poll_fields_other': poll_fields_other,
        'username': request.user.username,
        'lat':lat,
        'lon':lon,
        'incidence_video':incidence_video,
        'images':images,
        'user_create_incident':user_create_incident,
        'incidence_audio':incidence_audio,
        'poll_request_state':poll_request_state,
        'fields_record':fields_record,
        'flow': flow
    }
    
    return render(request, template_name, context)

@login_required
def department_view_read_only(request, request_id):
    session = int(check_profile_department(request))
    if session == 0:
        session = int(check_profile_management(request))
    if session == 0:
        session = int(check_profile_admin(request))
    if session == 0:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    
    flow = type_flow(request)
    try:
        poll_request = Request.objects.get(pk=request_id)
        poll_id = poll_request.poll_id
        poll_request_state = poll_request.request_state
    except Request.DoesNotExist:
        messages.add_message(request, messages.INFO, 'Hubo un error al acceder a la solicitud')
        return redirect('management_main')
    
    # Obtener los campos estándar activos
    poll_fields_standard = Fields.objects.filter(poll_id=poll_id, state='Activo', kind_field='standard').order_by('id')
    poll_fields_other = Fields.objects.filter(poll_id=poll_id, state='bloqueado', kind_field='other').order_by('id')
    
    # Verificar que existen campos para el poll_id dado
    if not poll_fields_standard.exists():
        messages.add_message(request, messages.INFO, 'No se encontraron campos para la encuesta proporcionada')
        return redirect('management_main')
    
    # Obtener la encuesta creada
    poll_data = Poll.objects.get(pk=poll_id)
    
    # Obtener las respuestas de las solicitudes que coincidan con los campos rescatados
    #request_answers = RequestAnswer.objects.filter(fields__in=poll_fields_standard).select_related('fields', 'request', 'user')
    request_answers = RequestAnswer.objects.filter(fields__in=poll_fields_standard).filter(request_id=request_id).select_related('fields', 'request', 'user')
    # Crear una lista de tuplas (nombre del campo, valor de respuesta)
    field_answer_pairs = []
    images = []
    lat = ''
    lon = ''
    incidence_video = ''
    incidence_audio = ''
    for field in poll_fields_standard:
        answer = request_answers.filter(fields=field).first()
        #field_answer_pairs.append((field.name, answer.request_answer_text if answer else ''))
        field_answer_pairs.append((field.name, answer.request_answer_text if answer else ''))
        if field.name == 'incidence_latitud':
            lat = answer.request_answer_text if answer else ''
        if field.name == 'incidence_longitud':
            lon = answer.request_answer_text if answer else ''
        if field.name == 'incidence_video':
            incidence_video = answer.request_answer_text if answer else ''
        if field.name == 'incidence_audio':
            incidence_audio = answer.request_answer_text if answer else ''

    for field in poll_fields_standard:
        if field.name == 'incidence_image':
            answer = request_answers.filter(fields=field)
            for a in answer:
                incidence_image = a.request_answer_text if answer else ''
                images.append(a.request_answer_text)

    brigades = Brigade.objects.select_related('user').filter(state='Activa')
    #usuario que genera la incidencia
    user_create_incident = poll_request.user.first_name+' '+poll_request.user.last_name

    # Preparar la lista con los datos necesarios
    brigades_data = []
    for brigade in brigades:
        user = brigade.user
        brigades_data.append({
            'brigade_id': brigade.id,
            'user_id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name
        })
    fields_record= RequestRecord.objects.filter(request_id=request_id)
    
    template_name = 'department_manager/department_view_read_only.html'
    
    context = {
        'brigades_data': brigades_data,
        'request_id': request_id,
        'poll_data': poll_data,
        'field_answer_pairs': field_answer_pairs,
        'poll_fields_other': poll_fields_other,
        'username': request.user.username,
        'lat':lat,
        'lon':lon,
        'incidence_video':incidence_video,
        'images':images,
        'user_create_incident':user_create_incident,
        'incidence_audio':incidence_audio,
        'poll_request_state':poll_request_state,
        'fields_record':fields_record,
        'flow': flow
    }
    
    return render(request, template_name, context)

@login_required
def department_view_read_only_admin(request, request_id):
    session = int(check_profile_department(request))
    if session == 0:
        session = int(check_profile_management(request))
    if session == 0:
        session = int(check_profile_admin(request))
    if session == 0:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    
    flow = type_flow(request)
    try:
        poll_request = Request.objects.get(pk=request_id)
        poll_id = poll_request.poll_id
        poll_request_state = poll_request.request_state
    except Request.DoesNotExist:
        messages.add_message(request, messages.INFO, 'Hubo un error al acceder a la solicitud')
        return redirect('management_main')
    
    # Obtener los campos estándar activos
    poll_fields_standard = Fields.objects.filter(poll_id=poll_id, state='Activo', kind_field='standard').order_by('id')
    poll_fields_other = Fields.objects.filter(poll_id=poll_id, state='bloqueado', kind_field='other').order_by('id')
    
    # Verificar que existen campos para el poll_id dado
    if not poll_fields_standard.exists():
        messages.add_message(request, messages.INFO, 'No se encontraron campos para la encuesta proporcionada')
        return redirect('management_main')
    
    # Obtener la encuesta creada
    poll_data = Poll.objects.get(pk=poll_id)
    
    # Obtener las respuestas de las solicitudes que coincidan con los campos rescatados
    #request_answers = RequestAnswer.objects.filter(fields__in=poll_fields_standard).select_related('fields', 'request', 'user')
    request_answers = RequestAnswer.objects.filter(fields__in=poll_fields_standard).filter(request_id=request_id).select_related('fields', 'request', 'user')

    # Crear una lista de tuplas (nombre del campo, valor de respuesta)
    field_answer_pairs = []
    images = []
    lat = ''
    lon = ''
    incidence_video = ''
    incidence_audio = ''
    for field in poll_fields_standard:
        answer = request_answers.filter(fields=field).first()
        #field_answer_pairs.append((field.name, answer.request_answer_text if answer else ''))
        field_answer_pairs.append((field.name, answer.request_answer_text if answer else ''))
        if field.name == 'incidence_latitud':
            lat = answer.request_answer_text if answer else ''
        if field.name == 'incidence_longitud':
            lon = answer.request_answer_text if answer else ''
        if field.name == 'incidence_video':
            incidence_video = answer.request_answer_text if answer else ''
        if field.name == 'incidence_audio':
            incidence_audio = answer.request_answer_text if answer else ''

    for field in poll_fields_standard:
        if field.name == 'incidence_image':
            answer = request_answers.filter(fields=field)
            for a in answer:
                incidence_image = a.request_answer_text if answer else ''
                images.append(incidence_image)

    brigades = Brigade.objects.select_related('user').filter(state='Activa')
    #usuario que genera la incidencia
    user_create_incident = poll_request.user.first_name+' '+poll_request.user.last_name

    # Preparar la lista con los datos necesarios
    brigades_data = []
    for brigade in brigades:
        user = brigade.user
        brigades_data.append({
            'brigade_id': brigade.id,
            'user_id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name
        })
    fields_record= RequestRecord.objects.filter(request_id=request_id)
    template_name = 'department_manager/department_view_read_only_admin.html'
    context = {
        'brigades_data': brigades_data,
        'request_id': request_id,
        'poll_data': poll_data,
        'field_answer_pairs': field_answer_pairs,
        'poll_fields_other': poll_fields_other,
        'username': request.user.username,
        'lat':lat,
        'lon':lon,
        'incidence_video':incidence_video,
        'images':images,
        'user_create_incident':user_create_incident,
        'incidence_audio':incidence_audio,
        'poll_request_state':poll_request_state,
        'fields_record':fields_record,
        'flow': flow
    }


    return render(request, template_name, context)

#Funcion para Encuestas ver Perfil
@login_required
def department_view_profile(request):
    session = int(check_profile_department(request))
    if session == 0:
        session = int(check_profile_management(request))
    if session == 0:
        messages.error(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    user = request.user
    profiles = Profile.objects.get(user_id = request.user.id)
    try:
        if request.method == 'POST':
            # Si se envía un formulario de edición, actualiza los datos
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST.get('email')
            user.save()
            messages.success(request, 'Perfil actualizado exitosamente.')
    except:
        messages.error(request, 'Hubo un error, favor contactese con los administradores')
        return redirect('check_group_main') 
    # Obtiene los datos actualizados del usuario
    user.refresh_from_db()

    context = {
        'user_data': user,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'email': user.email,
        'cargo': profiles.group.name, 
    }
    
    template_name = 'department_manager/department_view_profile.html'
    return render(request, template_name, context )

@csrf_exempt
@login_required
def department_derivar(request):
    if request.method == 'POST':
        # Obtener los datos del cuerpo de la solicitud JSON
        data = json.loads(request.body.decode('utf-8'))
        request_id = data.get('request_id')
        reason_for_delivery = data.get('reason_for_delivery')
        brigade_id = data.get('brigade_id')
        # Obtener el objeto Request utilizando el poll_id
        poll_request = get_object_or_404(Request, pk=request_id)

        # Actualizar el campo request_state a 'Derivada'
        poll_request.request_state = 'Derivada'
        poll_request.brigade_id = brigade_id
        poll_request.request_delivery = datetime.now().date()
        poll_request.save()
        #Agregamos registro razon derivar
        request_record_save = RequestRecord(
            user_id = request.user.id,
            request_id = request_id,
            request_record_kind = 'Derivada',
            request_record_text = reason_for_delivery,
            )
        request_record_save.save()        
        return JsonResponse({'status': 'success'}, status=200)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)
    
@csrf_exempt
@login_required
def department_cancelar(request):
    if request.method == 'POST':
        # Obtener los datos del cuerpo de la solicitud JSON
        data = json.loads(request.body.decode('utf-8'))
        request_id = data.get('request_id')
        reason_for_rejection = data.get('reason_for_rejection')
        # Obtener el objeto Request utilizando el poll_id
        poll_request = get_object_or_404(Request, pk=request_id)
        # Actualizar el campo request_state a 'Cancelada'
        poll_request.request_state = 'Cancelada'
        poll_request.save()
        #Agregamos registro razon cancelación
        request_record_save = RequestRecord(
            user_id = request.user.id,
            request_id = request_id,
            request_record_kind = 'Cancela',
            request_record_text = reason_for_rejection,
            )
        request_record_save.save()
        return JsonResponse({'status': 'success'}, status=200)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


#Funcion para manipular solicitudes
@login_required
def aceptar_solicitud(request):
    session = int(check_profile_department(request))
    if session == 0:
        session = int(check_profile_management(request))
    if session == 0:
        messages.error(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    try:
        if request.method == 'POST':
            request_open_list = Request.objects.filter(request_state='Abierta')
            return render(request, 'department_manager/department_main.html', {'request_open_list': request_open_list})
        else:
            # Manejar el caso en el que no se haya realizado una solicitud POST
            return HttpResponse('No se realizó una solicitud POST')
    except:
        messages.error(request, 'Hubo un error, favor contactese con los administradores')
        return redirect('check_group_main') 
    
#Funciones de usuario Direccion


#Funcion para Encuestas Enviadas
@login_required
def management_main(request):  
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    manuals_list = Manuals.objects.filter(manual_name='Manual Dirección').first
    request_open_list_array = Request.objects.filter(request_state='Iniciada').order_by('-id')
    request_open_list = []
    for re in request_open_list_array:
        poll_data =  Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        request_territorial_name = re.user.first_name+' '+re.user.last_name
        request_open_list.append({'id':re.id,'request_name':re.request_name,'request_date':re.request_date,'request_type':request_type,'request_territorial_name':request_territorial_name})
    template_name = 'department_manager/management_main.html'  



    return render(request,template_name, {'request_open_list': request_open_list,'manuals_list': manuals_list})

#Funcion para ver lista de Solicitudes Derivadas
@login_required
def management_list_derived(request,page=None):  
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    manuals_list = Manuals.objects.filter(manual_name='Manual Dirección').first
    request_delivery_list_array = Request.objects.filter(request_state='Derivada').order_by('-request_delivery')
    request_delivery_list = []
    for re in request_delivery_list_array:
        poll_data =  Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        #Cuadrialla a cargo
        user_brigade_incident_incharge = re.brigade.user.first_name+' '+re.brigade.user.last_name
        request_delivery_list.append({'id':re.id,'request_name':re.request_name,'request_delivery':re.request_delivery,'request_type':request_type,'user_brigade_incident_incharge':user_brigade_incident_incharge})    
    
    template_name = 'department_manager/management_list_derived.html'
    return render(request,template_name, {'request_delivery_list': request_delivery_list,'manuals_list': manuals_list})

#Funcion para ver lista de Solicitudes En Proceso
@login_required
def management_in_progress(request,page=None):  
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    manuals_list = Manuals.objects.filter(manual_name='Manual Dirección').first
    request_in_progress_list_array = Request.objects.filter(request_state='Proceso').order_by('-request_delivery')
    request_in_progress_list = []
    for re in request_in_progress_list_array:
        poll_data =  Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        #Cuadrialla a cargo
        user_brigade_incident_incharge = re.brigade.user.first_name+' '+re.brigade.user.last_name
        request_in_progress_list.append({'id':re.id,'request_name':re.request_name,'request_delivery':re.request_delivery,'request_accept':re.request_accept,'request_type':request_type,'user_brigade_incident_incharge':user_brigade_incident_incharge})    
    template_name = 'department_manager/management_in_progress.html'  
    return render(request,template_name, {'request_in_progress_list': request_in_progress_list,'manuals_list': manuals_list})

#Funcion para ver lista de Solicitudes Finalizadas
@login_required
def management_finish(request,page=None):  
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    manuals_list = Manuals.objects.filter(manual_name='Manual Dirección').first
    request_finish_list_array = Request.objects.filter(request_state='Finalizada').order_by('-request_delivery')
    request_in_finish_list = []
    for re in request_finish_list_array:
        poll_data =  Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        #Cuadrialla a cargo
        user_brigade_incident_incharge = re.brigade.user.first_name+' '+re.brigade.user.last_name
        request_in_finish_list.append({'id':re.id,'request_name':re.request_name,'request_delivery':re.request_delivery,'request_accept':re.request_accept,'request_accept':re.request_accept,'request_type':request_type,'user_brigade_incident_incharge':user_brigade_incident_incharge})    

    template_name = 'department_manager/management_finish.html'  
    return render(request,template_name, {'request_in_finish_list': request_in_finish_list,'manuals_list': manuals_list})

#Funcion para ver lista de Solicitudes Cerradas
@login_required
def management_list_closed(request,page=None):  
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    manuals_list = Manuals.objects.filter(manual_name='Manual Dirección').first
    request_close_list_array = Request.objects.filter(request_state='Cerrada').order_by('-request_delivery')
    request_closed_list = []
    for re in request_close_list_array:
        poll_data =  Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        #Cuadrialla a cargo
        user_brigade_incident_incharge = re.brigade.user.first_name+' '+re.brigade.user.last_name
        request_closed_list.append({'id':re.id,'request_name':re.request_name,'request_delivery':re.request_delivery,'request_accept':re.request_accept,'request_accept':re.request_accept,'request_close':re.request_close,'request_type':request_type,'user_brigade_incident_incharge':user_brigade_incident_incharge})    

    template_name = 'department_manager/management_list_closed.html'  
    return render(request,template_name, {'request_closed_list': request_closed_list,'manuals_list': manuals_list})

#Funcion para ver Encuestas Enviadas
@login_required
def management_view(request, request_id):
    session = int(check_profile_management(request))
    if session == 0:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')

    flow = type_flow(request)
    try:
        poll_request = Request.objects.get(pk=request_id)
        poll_id = poll_request.poll_id
        poll_request_state = poll_request.request_state
    except Request.DoesNotExist:
        messages.add_message(request, messages.INFO, 'Hubo un error al acceder a la solicitud')
        return redirect('management_main')
    
    # Obtener los campos estándar activos
    poll_fields_standard = Fields.objects.filter(poll_id=poll_id, state='Activo', kind_field='standard').order_by('id')
    poll_fields_other = Fields.objects.filter(poll_id=poll_id, state='bloqueado', kind_field='other').order_by('id')
    
    # Verificar que existen campos para el poll_id dado
    if not poll_fields_standard.exists():
        messages.add_message(request, messages.INFO, 'No se encontraron campos para la encuesta proporcionada')
        return redirect('management_main')
    
    # Obtener la encuesta creada
    poll_data = Poll.objects.get(pk=poll_id)
    
    # Obtener las respuestas de las solicitudes que coincidan con los campos rescatados
    request_answers = RequestAnswer.objects.filter(fields__in=poll_fields_standard).filter(request_id=request_id).select_related('fields', 'request', 'user')
    
    # Crear una lista de tuplas (nombre del campo, valor de respuesta)
    field_answer_pairs = []
    images = []
    lat = ''
    lon = ''
    incidence_video = ''
    incidence_audio = ''
    for field in poll_fields_standard:
        answer = request_answers.filter(fields=field).first()
        #field_answer_pairs.append((field.name, answer.request_answer_text if answer else ''))
        field_answer_pairs.append((field.name, answer.request_answer_text if answer else ''))
        if field.name == 'incidence_latitud':
            lat = answer.request_answer_text if answer else ''
        if field.name == 'incidence_longitud':
            lon = answer.request_answer_text if answer else ''
        if field.name == 'incidence_video':
            incidence_video = answer.request_answer_text if answer else ''
        if field.name == 'incidence_audio':
            incidence_audio = answer.request_answer_text if answer else ''

    for field in poll_fields_standard:
        if field.name == 'incidence_image':
            answer = request_answers.filter(fields=field)
            for a in answer:
                incidence_image = a.request_answer_text if answer else ''
                images.append(incidence_image)

            
    brigades = Brigade.objects.select_related('user').filter(state='Activa')
    #usuario que genera la incidencia
    user_create_incident = poll_request.user.first_name+' '+poll_request.user.last_name

    # Preparar la lista con los datos necesarios
    brigades_data = []
    for brigade in brigades:
        user = brigade.user
        brigades_data.append({
            'brigade_id': brigade.id,
            'user_id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name
        })
    fields_record= RequestRecord.objects.filter(request_id=request_id)
    template_name = 'department_manager/management_view.html'
    context = {
        'brigades_data': brigades_data,
        'request_id': request_id,
        'poll_data': poll_data,
        'field_answer_pairs': field_answer_pairs,
        'poll_fields_other': poll_fields_other,
        'username': request.user.username,
        'lat':lat,
        'lon':lon,
        'incidence_video':incidence_video,
        'images':images,
        'user_create_incident':user_create_incident,
        'incidence_audio':incidence_audio,
        'poll_request_state':poll_request_state,
        'fields_record':fields_record,
        'flow': flow
    }
    return render(request, template_name, context)

@login_required
def management_view_read_only(request, request_id):
    session = int(check_profile_management(request))
    if session == 0:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    
    flow = type_flow(request)
    try:
        poll_request = Request.objects.get(pk=request_id)
        poll_id = poll_request.poll_id
        poll_request_state = poll_request.request_state        
    except Request.DoesNotExist:
        messages.add_message(request, messages.INFO, 'Hubo un error al acceder a la solicitud')
        return redirect('management_main')
    
    # Obtener los campos estándar activos
    poll_fields_standard = Fields.objects.filter(poll_id=poll_id, state='Activo', kind_field='standard').order_by('id')
    poll_fields_other = Fields.objects.filter(poll_id=poll_id, state='bloqueado', kind_field='other').order_by('id')
    
    # Verificar que existen campos para el poll_id dado
    if not poll_fields_standard.exists():
        messages.add_message(request, messages.INFO, 'No se encontraron campos para la encuesta proporcionada')
        return redirect('management_main')
    
    # Obtener la encuesta creada
    poll_data = Poll.objects.get(pk=poll_id)
    
    # Obtener las respuestas de las solicitudes que coincidan con los campos rescatados
    request_answers = RequestAnswer.objects.filter(fields__in=poll_fields_standard).select_related('fields', 'request', 'user')
    
    # Crear una lista de tuplas (nombre del campo, valor de respuesta)
    field_answer_pairs = []
    images = []
    lat = ''
    lon = ''
    incidence_video = ''
    incidence_audio = ''
    for field in poll_fields_standard:
        answer = request_answers.filter(fields=field).first()
        #field_answer_pairs.append((field.name, answer.request_answer_text if answer else ''))
        field_answer_pairs.append((field.name, answer.request_answer_text if answer else ''))
        if field.name == 'incidence_latitud':
            lat = answer.request_answer_text if answer else ''
        if field.name == 'incidence_longitud':
            lon = answer.request_answer_text if answer else ''
        if field.name == 'incidence_video':
            incidence_video = answer.request_answer_text if answer else ''
        if field.name == 'incidence_audio':
            incidence_audio = answer.request_answer_text if answer else ''

    for field in poll_fields_standard:
        if field.name == 'incidence_image':
            answer = request_answers.filter(fields=field)
            for a in answer:
                incidence_image = a.request_answer_text if answer else ''
                images.append(incidence_image)

            
    brigades = Brigade.objects.select_related('user').filter(state='Activa')
    #usuario que genera la incidencia
    user_create_incident = poll_request.user.first_name+' '+poll_request.user.last_name

    # Preparar la lista con los datos necesarios
    brigades_data = []
    for brigade in brigades:
        user = brigade.user
        brigades_data.append({
            'brigade_id': brigade.id,
            'user_id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name
        })
    fields_record= RequestRecord.objects.filter(request_id=request_id)
    
    template_name = 'department_manager/management_view_read_only.html'
    
    context = {
        'brigades_data': brigades_data,
        'request_id': request_id,
        'poll_data': poll_data,
        'field_answer_pairs': field_answer_pairs,
        'poll_fields_other': poll_fields_other,
        'username': request.user.username,
        'lat':lat,
        'lon':lon,
        'incidence_video':incidence_video,
        'images':images,
        'user_create_incident':user_create_incident,
        'incidence_audio':incidence_audio,
        'poll_request_state':poll_request_state,
        'fields_record':fields_record,
        'flow': flow
    }
    
    return render(request, template_name, context)

#Funcion para Encuestas ver Perfil
@login_required
def management_view_profile(request):
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('logout')
    user = request.user
    profiles = Profile.objects.get(user_id = request.user.id)

    try:
        if request.method == 'POST':
            # Si se envía un formulario de edición, actualiza los datos
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST.get('email')
            user.save()
            messages.success(request, 'Perfil actualizado exitosamente.')
    except:
        messages.error(request, 'Hubo un error, favor contactese con los administradores')
        return redirect('check_group_main')

    # Obtiene los datos actualizados del usuario
    user.refresh_from_db()

    context = {
        'user_data': user,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'email': user.email,
        'cargo': profiles.group.name, 
    }
    
    template_name = 'department_manager/management_view_profile.html'
    return render(request, template_name, context )

@csrf_exempt
@login_required
def management_derivar(request):
    if request.method == 'POST':
        # Obtener los datos del cuerpo de la solicitud JSON
        data = json.loads(request.body.decode('utf-8'))
        request_id = data.get('request_id')  
        reason_for_delivery = data.get('reason_for_delivery')
        brigade_id = data.get('brigade_id')              
        # Obtener el objeto Request utilizando el poll_id
        poll_request = get_object_or_404(Request, pk=request_id)

        # Actualizar el campo request_state a 'Derivada'
        poll_request.request_state = 'Derivada'
        poll_request.brigade_id = data.get('cuadrilla')
        poll_request.request_delivery = datetime.now().date()        
        poll_request.save()
        #Agregamos registro razon derivar
        request_record_save = RequestRecord(
            user_id = request.user.id,
            request_id = request_id,
            request_record_kind = 'Derivada',
            request_record_text = reason_for_delivery,
            )
        request_record_save.save()   

        return JsonResponse({'status': 'success'}, status=200)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)
    
@csrf_exempt
@login_required
def management_cancelar(request):
    if request.method == 'POST':
        # Obtener los datos del cuerpo de la solicitud JSON
        data = json.loads(request.body.decode('utf-8'))
        request_id = data.get('request_id')
        reason_for_rejection = data.get('reason_for_rejection')

        # Obtener el objeto Request utilizando el poll_id
        poll_request = get_object_or_404(Request, pk=request_id)


        # Actualizar el campo request_state a 'Cancelada'
        poll_request.request_state = 'Cancelada'
        poll_request.save()
        #Agregamos registro razon cancelación
        request_record_save = RequestRecord(
            user_id = request.user.id,
            request_id = request_id,
            request_record_kind = 'Cancela',
            request_record_text = reason_for_rejection,
            )
        request_record_save.save()
        return JsonResponse({'status': 'success'}, status=200)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)
    

@login_required
#Reporte Excel de Solicitudes abiertas
def department_report_list(request,page=None):  
    try:
        session = int(check_profile_department(request))
        if session == 0:
            messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
            return redirect('logout')
    
        list_array = Request.objects.filter(request_state = 'Iniciada').order_by('request_name')
        response = HttpResponse(content_type = 'application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ReporteEncuestasIniciadas.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Encuestas')
        row_num = 0
        columns = ['Nombre', 'Tipo incidencia', 'Solicitada por', 'Fecha','Observaciones','Latitud','Longitud']
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        
        font_style = xlwt.XFStyle()
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'dd/MM/yyyy'

        ws.col(0).width = 256 * 15
        ws.col(1).width = 256 * 30
        ws.col(2).width = 256 * 20
        ws.col(3).width = 256 * 15
        ws.col(4).width = 256 * 15
        ws.col(5).width = 256 * 15        
        ws.col(6).width = 256 * 15

        for request in list_array:
            #obtenemos los id de las preguntas
            poll_id = request.poll_id
            #buscamos los id de preguntas (latitud,longitud,observacion)
            request_answer_incidence_description = ''
            poll_fields_incidence_description_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_description').filter(state='Activo').count()
            if poll_fields_incidence_description_count > 0:
                poll_fields_incidence_description_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_description').filter(state='Activo').get()
                poll_fields_incidence_description_id = poll_fields_incidence_description_data.id 
                #buscamos la información
                request_answer_incidence_description_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_description_id).count()
                if request_answer_incidence_description_count > 0:
                    request_answer_incidence_description_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_description_id).get()
                    request_answer_incidence_description = request_answer_incidence_description_data.request_answer_text

            request_answer_incidence_latitud = ''
            poll_fields_incidence_latitud_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_latitud').filter(state='Activo').count()
            if poll_fields_incidence_latitud_count > 0:
                poll_fields_incidence_latitud_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_latitud').filter(state='Activo').get()
                poll_fields_incidence_latitud_id = poll_fields_incidence_latitud_data.id 
                #buscamos la información
                request_answer_incidence_latitud_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_latitud_id).count()
                if request_answer_incidence_latitud_count > 0:
                    request_answer_incidence_latitud_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_latitud_id).get()
                    request_answer_incidence_latitud = request_answer_incidence_latitud_data.request_answer_text

            request_answer_incidence_longitud = ''
            poll_fields_incidence_longitud_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_longitud').filter(state='Activo').count()
            if poll_fields_incidence_longitud_count > 0:
                poll_fields_incidence_longitud_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_longitud').filter(state='Activo').get()
                poll_fields_incidence_longitud_id = poll_fields_incidence_longitud_data.id 
                #buscamos la información
                request_answer_incidence_longitud_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_longitud_id).count()
                if request_answer_incidence_longitud_count > 0:
                    request_answer_incidence_longitud_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_longitud_id).get()
                    request_answer_incidence_longitud = request_answer_incidence_longitud_data.request_answer_text
            
            row_num += 1
            ws.write(row_num, 0, request.request_name, font_style)
            ws.write(row_num, 1, request.poll.incident.name, font_style)
            ws.write(row_num, 2, request.user.first_name +' '+ request.user.last_name, font_style)
            ws.write(row_num, 3, request.request_date, date_format)
            ws.write(row_num, 4, request_answer_incidence_description, font_style)
            ws.write(row_num, 5, request_answer_incidence_latitud, font_style)
            ws.write(row_num, 6, request_answer_incidence_longitud, font_style)


        wb.save(response)
        return response
    
    except Exception as e:
        messages.add_message(request, messages.INFO, f'Error al generar el reporte: {str(e)}')
        return redirect('deparment_main')


@login_required
#Reporte Excel de Solicitudes derivadas
def department_report_list_derived(request,page=None):  
    try:
        session = int(check_profile_department(request))
        if session == 0:
            messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
            return redirect('logout')
    
        list_array = Request.objects.filter(request_state = 'Derivada').order_by('request_name')
        response = HttpResponse(content_type = 'application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ReporteEncuestasDerivadas.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Encuestas')
        row_num = 0
        columns = ['Nombre', 'Tipo incidencia', 'Solicitada por', 'Fecha Derivación', 'Resolutor Asignado','Observaciones','Latitud','Longitud']
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        
        font_style = xlwt.XFStyle()
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'dd/MM/yyyy'

        ws.col(0).width = 256 * 15
        ws.col(1).width = 256 * 30
        ws.col(2).width = 256 * 20
        ws.col(3).width = 256 * 15
        ws.col(4).width = 256 * 20
        ws.col(5).width = 256 * 15
        ws.col(6).width = 256 * 15        
        ws.col(7).width = 256 * 15

        for request in list_array:
            #obtenemos los id de las preguntas
            poll_id = request.poll_id
            #buscamos los id de preguntas (latitud,longitud,observacion)
            request_answer_incidence_description = ''
            poll_fields_incidence_description_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_description').filter(state='Activo').count()
            if poll_fields_incidence_description_count > 0:
                poll_fields_incidence_description_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_description').filter(state='Activo').get()
                poll_fields_incidence_description_id = poll_fields_incidence_description_data.id 
                #buscamos la información
                request_answer_incidence_description_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_description_id).count()
                if request_answer_incidence_description_count > 0:
                    request_answer_incidence_description_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_description_id).get()
                    request_answer_incidence_description = request_answer_incidence_description_data.request_answer_text

            request_answer_incidence_latitud = ''
            poll_fields_incidence_latitud_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_latitud').filter(state='Activo').count()
            if poll_fields_incidence_latitud_count > 0:
                poll_fields_incidence_latitud_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_latitud').filter(state='Activo').get()
                poll_fields_incidence_latitud_id = poll_fields_incidence_latitud_data.id 
                #buscamos la información
                request_answer_incidence_latitud_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_latitud_id).count()
                if request_answer_incidence_latitud_count > 0:
                    request_answer_incidence_latitud_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_latitud_id).get()
                    request_answer_incidence_latitud = request_answer_incidence_latitud_data.request_answer_text

            request_answer_incidence_longitud = ''
            poll_fields_incidence_longitud_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_longitud').filter(state='Activo').count()
            if poll_fields_incidence_longitud_count > 0:
                poll_fields_incidence_longitud_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_longitud').filter(state='Activo').get()
                poll_fields_incidence_longitud_id = poll_fields_incidence_longitud_data.id 
                #buscamos la información
                request_answer_incidence_longitud_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_longitud_id).count()
                if request_answer_incidence_longitud_count > 0:
                    request_answer_incidence_longitud_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_longitud_id).get()
                    request_answer_incidence_longitud = request_answer_incidence_longitud_data.request_answer_text

            row_num += 1
            ws.write(row_num, 0, request.request_name, font_style)
            ws.write(row_num, 1, request.poll.incident.name, font_style)
            ws.write(row_num, 2, request.user.first_name +' '+ request.user.last_name, font_style)
            ws.write(row_num, 3, request.request_delivery, date_format)
            ws.write(row_num, 4, request.brigade.user.first_name +' '+ request.brigade.user.last_name, font_style)
            ws.write(row_num, 5, request_answer_incidence_description, font_style)
            ws.write(row_num, 6, request_answer_incidence_latitud, font_style)
            ws.write(row_num, 7, request_answer_incidence_longitud, font_style)

        wb.save(response)
        return response
    
    except Exception as e:
        messages.add_message(request, messages.INFO, f'Error al generar el reporte: {str(e)}')
        return redirect('deparment_list_derived')
    

@login_required
#Reporte Excel de Solicitudes en Proceso
def department_report_list_progress(request,page=None):  
    try:
        session = int(check_profile_department(request))
        if session == 0:
            messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
            return redirect('logout')
    
        list_array = Request.objects.filter(request_state = 'Proceso').order_by('request_name')
        response = HttpResponse(content_type = 'application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ReporteEncuestasenProceso.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Encuestas')
        row_num = 0
        columns = ['Nombre', 'Tipo incidencia', 'Fecha Derivación', 'Fecha Inicio', 'Resolutor Asignado','Observaciones','Latitud','Longitud']
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        
        font_style = xlwt.XFStyle()
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'dd/MM/yyyy'

        ws.col(0).width = 256 * 15
        ws.col(1).width = 256 * 30
        ws.col(2).width = 256 * 20
        ws.col(3).width = 256 * 15
        ws.col(4).width = 256 * 20
        ws.col(5).width = 256 * 15
        ws.col(6).width = 256 * 15        
        ws.col(7).width = 256 * 15

        for request in list_array:
            #obtenemos los id de las preguntas
            poll_id = request.poll_id
            #buscamos los id de preguntas (latitud,longitud,observacion)
            request_answer_incidence_description = ''
            poll_fields_incidence_description_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_description').filter(state='Activo').count()
            if poll_fields_incidence_description_count > 0:
                poll_fields_incidence_description_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_description').filter(state='Activo').get()
                poll_fields_incidence_description_id = poll_fields_incidence_description_data.id 
                #buscamos la información
                request_answer_incidence_description_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_description_id).count()
                if request_answer_incidence_description_count > 0:
                    request_answer_incidence_description_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_description_id).get()
                    request_answer_incidence_description = request_answer_incidence_description_data.request_answer_text

            request_answer_incidence_latitud = ''
            poll_fields_incidence_latitud_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_latitud').filter(state='Activo').count()
            if poll_fields_incidence_latitud_count > 0:
                poll_fields_incidence_latitud_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_latitud').filter(state='Activo').get()
                poll_fields_incidence_latitud_id = poll_fields_incidence_latitud_data.id 
                #buscamos la información
                request_answer_incidence_latitud_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_latitud_id).count()
                if request_answer_incidence_latitud_count > 0:
                    request_answer_incidence_latitud_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_latitud_id).get()
                    request_answer_incidence_latitud = request_answer_incidence_latitud_data.request_answer_text

            request_answer_incidence_longitud = ''
            poll_fields_incidence_longitud_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_longitud').filter(state='Activo').count()
            if poll_fields_incidence_longitud_count > 0:
                poll_fields_incidence_longitud_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_longitud').filter(state='Activo').get()
                poll_fields_incidence_longitud_id = poll_fields_incidence_longitud_data.id 
                #buscamos la información
                request_answer_incidence_longitud_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_longitud_id).count()
                if request_answer_incidence_longitud_count > 0:
                    request_answer_incidence_longitud_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_longitud_id).get()
                    request_answer_incidence_longitud = request_answer_incidence_longitud_data.request_answer_text
            
            row_num += 1
            ws.write(row_num, 0, request.request_name, font_style)
            ws.write(row_num, 1, request.poll.incident.name, font_style)
            ws.write(row_num, 2, request.request_delivery, date_format)
            ws.write(row_num, 3, request.request_accept, date_format)
            ws.write(row_num, 4, request.brigade.user.first_name +' '+ request.brigade.user.last_name, font_style)
            ws.write(row_num, 5, request_answer_incidence_description, font_style)
            ws.write(row_num, 6, request_answer_incidence_latitud, font_style)
            ws.write(row_num, 7, request_answer_incidence_longitud, font_style)

        wb.save(response)
        return response
    
    except Exception as e:
        messages.add_message(request, messages.INFO, f'Error al generar el reporte: {str(e)}')
        return redirect('deparment_in_progress')
    
@login_required
#Reporte Excel de Solicitudes Finalizadas
def department_report_list_finish(request,page=None):  
    try:
        session = int(check_profile_department(request))
        if session == 0:
            messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
            return redirect('logout')
    
        list_array = Request.objects.filter(request_state = 'Finalizada').order_by('request_name')
        response = HttpResponse(content_type = 'application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ReporteEncuestasFinalizadas.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Encuestas')
        row_num = 0
        columns = ['Nombre', 'Tipo incidencia', 'Fecha Derivación', 'Fecha Inicio', 'Fecha Finalización', 'Resolutor Asignado','Observaciones','Latitud','Longitud']
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        
        font_style = xlwt.XFStyle()
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'dd/MM/yyyy'

        ws.col(0).width = 256 * 15
        ws.col(1).width = 256 * 30
        ws.col(2).width = 256 * 20
        ws.col(3).width = 256 * 15
        ws.col(4).width = 256 * 20
        ws.col(5).width = 256 * 20
        ws.col(6).width = 256 * 15
        ws.col(7).width = 256 * 15
        ws.col(8).width = 256 * 15        

        for request in list_array:
            #obtenemos los id de las preguntas
            poll_id = request.poll_id
            #buscamos los id de preguntas (latitud,longitud,observacion)
            request_answer_incidence_description = ''
            poll_fields_incidence_description_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_description').filter(state='Activo').count()
            if poll_fields_incidence_description_count > 0:
                poll_fields_incidence_description_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_description').filter(state='Activo').get()
                poll_fields_incidence_description_id = poll_fields_incidence_description_data.id 
                #buscamos la información
                request_answer_incidence_description_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_description_id).count()
                if request_answer_incidence_description_count > 0:
                    request_answer_incidence_description_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_description_id).get()
                    request_answer_incidence_description = request_answer_incidence_description_data.request_answer_text

            request_answer_incidence_latitud = ''
            poll_fields_incidence_latitud_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_latitud').filter(state='Activo').count()
            if poll_fields_incidence_latitud_count > 0:
                poll_fields_incidence_latitud_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_latitud').filter(state='Activo').get()
                poll_fields_incidence_latitud_id = poll_fields_incidence_latitud_data.id 
                #buscamos la información
                request_answer_incidence_latitud_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_latitud_id).count()
                if request_answer_incidence_latitud_count > 0:
                    request_answer_incidence_latitud_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_latitud_id).get()
                    request_answer_incidence_latitud = request_answer_incidence_latitud_data.request_answer_text

            request_answer_incidence_longitud = ''
            poll_fields_incidence_longitud_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_longitud').filter(state='Activo').count()
            if poll_fields_incidence_longitud_count > 0:
                poll_fields_incidence_longitud_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_longitud').filter(state='Activo').get()
                poll_fields_incidence_longitud_id = poll_fields_incidence_longitud_data.id 
                #buscamos la información
                request_answer_incidence_longitud_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_longitud_id).count()
                if request_answer_incidence_longitud_count > 0:
                    request_answer_incidence_longitud_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_longitud_id).get()
                    request_answer_incidence_longitud = request_answer_incidence_longitud_data.request_answer_text
            
            row_num += 1
            ws.write(row_num, 0, request.request_name, font_style)
            ws.write(row_num, 1, request.poll.incident.name, font_style)
            ws.write(row_num, 2, request.request_delivery, date_format)
            ws.write(row_num, 3, request.request_accept, date_format)
            ws.write(row_num, 4, request.request_finish, date_format)
            ws.write(row_num, 5, request.brigade.user.first_name +' '+ request.brigade.user.last_name, font_style)
            ws.write(row_num, 6, request_answer_incidence_description, font_style)
            ws.write(row_num, 7, request_answer_incidence_latitud, font_style)
            ws.write(row_num, 8, request_answer_incidence_longitud, font_style)

        wb.save(response)
        return response
    
    except Exception as e:
        messages.add_message(request, messages.INFO, f'Error al generar el reporte: {str(e)}')
        return redirect('deparment_finish')
    
@login_required
#Reporte Excel de Solicitudes Cerradas
def department_report_list_closed(request,page=None):  
    try:
        session = int(check_profile_department(request))
        if session == 0:
            messages.warning(request, 'Intenta ingresar a una area para la que no tiene permisos')
            return redirect('logout')
    
        list_array = Request.objects.filter(request_state = 'Cerrada').order_by('request_name')
        response = HttpResponse(content_type = 'application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ReporteEncuestasCerradas.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Encuestas')
        row_num = 0
        columns = ['Nombre', 'Tipo incidencia', 'Fecha Derivación', 'Fecha Inicio', 'Fecha Finalización', 'Fecha Cierre', 'Resolutor Asignado','Observaciones','Latitud','Longitud']
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        
        font_style = xlwt.XFStyle()
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'dd/MM/yyyy'

        ws.col(0).width = 256 * 15
        ws.col(1).width = 256 * 30
        ws.col(2).width = 256 * 20
        ws.col(3).width = 256 * 15
        ws.col(4).width = 256 * 20
        ws.col(5).width = 256 * 15
        ws.col(6).width = 256 * 20
        ws.col(7).width = 256 * 20
        ws.col(8).width = 256 * 15
        ws.col(9).width = 256 * 20

        for request in list_array:
            #obtenemos los id de las preguntas
            poll_id = request.poll_id
            #buscamos los id de preguntas (latitud,longitud,observacion)
            request_answer_incidence_description = ''
            poll_fields_incidence_description_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_description').filter(state='Activo').count()
            if poll_fields_incidence_description_count > 0:
                poll_fields_incidence_description_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_description').filter(state='Activo').get()
                poll_fields_incidence_description_id = poll_fields_incidence_description_data.id 
                #buscamos la información
                request_answer_incidence_description_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_description_id).count()
                if request_answer_incidence_description_count > 0:
                    request_answer_incidence_description_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_description_id).get()
                    request_answer_incidence_description = request_answer_incidence_description_data.request_answer_text

            request_answer_incidence_latitud = ''
            poll_fields_incidence_latitud_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_latitud').filter(state='Activo').count()
            if poll_fields_incidence_latitud_count > 0:
                poll_fields_incidence_latitud_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_latitud').filter(state='Activo').get()
                poll_fields_incidence_latitud_id = poll_fields_incidence_latitud_data.id 
                #buscamos la información
                request_answer_incidence_latitud_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_latitud_id).count()
                if request_answer_incidence_latitud_count > 0:
                    request_answer_incidence_latitud_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_latitud_id).get()
                    request_answer_incidence_latitud = request_answer_incidence_latitud_data.request_answer_text

            request_answer_incidence_longitud = ''
            poll_fields_incidence_longitud_count = Fields.objects.filter(poll_id=poll_id).filter(name='incidence_longitud').filter(state='Activo').count()
            if poll_fields_incidence_longitud_count > 0:
                poll_fields_incidence_longitud_data = Fields.objects.filter(poll_id=request.poll.id).filter(name='incidence_longitud').filter(state='Activo').get()
                poll_fields_incidence_longitud_id = poll_fields_incidence_longitud_data.id 
                #buscamos la información
                request_answer_incidence_longitud_count = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_longitud_id).count()
                if request_answer_incidence_longitud_count > 0:
                    request_answer_incidence_longitud_data = RequestAnswer.objects.filter(request_id=request.id).filter(fields_id=poll_fields_incidence_longitud_id).get()
                    request_answer_incidence_longitud = request_answer_incidence_longitud_data.request_answer_text
                    
            row_num += 1
            ws.write(row_num, 0, request.request_name, font_style)
            ws.write(row_num, 1, request.poll.incident.name, font_style)
            ws.write(row_num, 2, request.request_delivery, date_format)
            ws.write(row_num, 3, request.request_accept, date_format)
            ws.write(row_num, 4, request.request_finish, date_format)
            ws.write(row_num, 5, request.request_close, date_format)
            ws.write(row_num, 6, request.brigade.user.first_name +' '+ request.brigade.user.last_name, font_style)
            ws.write(row_num, 7, request_answer_incidence_description, font_style)
            ws.write(row_num, 8, request_answer_incidence_latitud, font_style)
            ws.write(row_num, 9, request_answer_incidence_longitud, font_style)

        wb.save(response)
        return response
    
    except Exception as e:
        messages.add_message(request, messages.INFO, f'Error al generar el reporte: {str(e)}')
        return redirect('deparment_finish')

#Función reporte excel activas
@login_required
def export_request_open_excel(request):
    # Validar sesión
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a un área para la que no tiene permisos')
        return redirect('logout')
    # Obtener solicitudes activas filtradas
    request_open_list_array = Request.objects.filter(request_state='Iniciada').order_by('-id')

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Listado Solicitudes Abiertas'

    # Encabezados del Excel
    ws.append(['Nombre Solicitud', 'Tipo de Incidente', 'Nombre Territorial', 'Fecha de Solicitud'])

    # Añadir datos al Excel
    for re in request_open_list_array:
        poll_data = Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        request_territorial_name = re.user.first_name + ' ' + re.user.last_name
        ws.append([re.request_name, request_type, request_territorial_name, re.request_date])

    # Crear respuesta para el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_Solicitudes_Abiertas.xlsx'
    wb.save(response)

    return response 

#Función reporte excel derivadas
@login_required
def export_request_delivery_excel(request):
    # Validar sesión
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a un área para la que no tienes permisos')
        return redirect('logout')
    
    # Obtener solicitudes derivadas filtradas
    request_delivery_list_array = Request.objects.filter(request_state='Derivada').order_by('-id')

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Listado Solicitudes Derivadas'

    # Encabezados del Excel
    ws.append(['Nombre', 'Tipo Incidencia', 'Fecha Derivación', 'Resolutor Asignado'])

    # Añadir datos al Excel
    for re in request_delivery_list_array:
        poll_data = Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        request_delivery_date = re.request_date
        user_brigade_incident_incharge = re.user.first_name + ' ' + re.user.last_name
        ws.append([re.request_name, request_type, request_delivery_date, user_brigade_incident_incharge])

    # Crear respuesta para el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_Solicitudes_Derivadas.xlsx'
    wb.save(response)

    return response

#Funcion reporte excel en progreso
@login_required
def export_request_in_progress_excel(request):
    # Validar sesión
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a un área para la que no tienes permisos')
        return redirect('logout')
    
    # Obtener solicitudes en proceso filtradas
    request_in_progress_list_array = Request.objects.filter(request_state='Proceso').order_by('-id')

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Listado Solicitudes en Proceso'

    # Encabezados del Excel
    ws.append(['Nombre', 'Tipo Incidencia', 'Fecha Derivación', 'Fecha Inicio', 'Resolutor Asignado'])

    # Añadir datos al Excel
    for re in request_in_progress_list_array:
        poll_data = Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        request_delivery_date = re.request_delivery  # Ajustado para el nombre del campo correcto
        request_accept_date = re.request_accept  # Asegúrate de que este campo existe en tu modelo
        user_brigade_incident_incharge = re.user.first_name + ' ' + re.user.last_name
        ws.append([re.request_name, request_type, request_delivery_date, request_accept_date, user_brigade_incident_incharge])

    # Crear respuesta para el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_Solicitudes_en_Proceso.xlsx'
    wb.save(response)

    return response


#Funcion reporte excel finalizadas
@login_required
def export_request_finalized_excel(request):
    # Validar sesión
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a un área para la que no tienes permisos')
        return redirect('logout')
    
    # Obtener solicitudes finalizadas filtradas
    request_finalized_list_array = Request.objects.filter(request_state='Finalizada').order_by('-id')

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Listado Solicitudes Finalizadas'

    # Encabezados del Excel
    ws.append(['Nombre', 'Tipo Incidencia', 'Fecha Derivación', 'Fecha Inicio', 'Fecha Finalización', 'Resolutor Asignado'])

    # Añadir datos al Excel
    for re in request_finalized_list_array:
        poll_data = Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        request_delivery_date = re.request_delivery  # Fecha de derivación
        request_accept_date = re.request_accept  # Fecha de inicio
        request_finish_date = re.request_finish  # Fecha de finalización
        user_brigade_incident_incharge = re.user.first_name + ' ' + re.user.last_name  # Resolutor asignado

        # Agregar fila de datos al Excel
        ws.append([re.request_name, request_type, request_delivery_date, request_accept_date, request_finish_date, user_brigade_incident_incharge])

    # Crear respuesta para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_Solicitudes_Finalizadas.xlsx'
    wb.save(response)

    return response

#Funcion reporte excel cerradas
@login_required
def export_request_closed_excel(request):
    # Validar sesión
    session = int(check_profile_management(request))
    if session == 0:
        messages.warning(request, 'Intenta ingresar a un área para la que no tienes permisos')
        return redirect('logout')

    # Obtener solicitudes cerradas filtradas
    request_closed_list_array = Request.objects.filter(request_state='Cerrada').order_by('-id')

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Listado Solicitudes Cerradas'

    # Encabezados del Excel
    ws.append(['Nombre', 'Tipo Incidencia', 'Fecha Derivación', 'Fecha Inicio', 'Fecha Finalización', 'Fecha Cierre', 'Resolutor Asignado'])

    # Añadir datos al Excel
    for re in request_closed_list_array:
        poll_data = Poll.objects.get(pk=re.poll_id)
        poll_data_incident_data = Incident.objects.get(pk=poll_data.incident_id)
        request_type = poll_data_incident_data.name
        request_delivery_date = re.request_delivery  # Fecha de derivación
        request_accept_date = re.request_accept  # Fecha de inicio
        request_finish_date = re.request_finish  # Fecha de finalización
        request_close_date = re.request_close  # Fecha de cierre
        user_brigade_incident_incharge = re.user.first_name + ' ' + re.user.last_name  # Resolutor asignado

        # Agregar fila de datos al Excel
        ws.append([re.request_name, request_type, request_delivery_date, request_accept_date, request_finish_date, request_close_date, user_brigade_incident_incharge])

    # Crear respuesta para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_Solicitudes_Cerradas.xlsx'
    wb.save(response)

    return response